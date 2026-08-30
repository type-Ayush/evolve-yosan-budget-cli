import argparse
import calendar
from datetime import datetime
import json
import os
from pathlib import Path
import shutil
import sqlite3
import sys
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from auth import clear_session, delete_account_flow, require_login, show_profile_view

# Enable Windows ANSI Virtual Terminal Processing
os.system("")

DATA_DIR = Path.home() / "Documents"
REPORTS_DIR = DATA_DIR / "yosan_reports"
CONFIG_FILE = DATA_DIR / ".yosan_config.json"


# ==========================================
# COLOR CONSTANTS
# ==========================================
class C:
    RESET   = "\033[0m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    CYAN    = "\033[38;2;0;210;255m"
    GREEN   = "\033[38;2;46;204;113m"
    RED     = "\033[38;2;231;76;60m"
    YELLOW  = "\033[38;2;241;196;15m"
    PURPLE  = "\033[38;2;165;94;234m"
    ORANGE  = "\033[38;2;250;130;49m"
    GRAY    = "\033[38;2;127;143;166m"
    WHITE   = "\033[38;2;245;246;250m"


CATEGORIES = {
    "-m": "Mess Food",
    "-c": "Clothes",
    "-a": "Accessories",
    "-v": "Savings",
}

CATEGORY_ANSI = {
    "Mess Food": "\033[38;2;52;152;219m",    # Vibrant Blue
    "Clothes": "\033[38;2;46;204;113m",      # Emerald Green
    "Accessories": "\033[38;2;243;156;18m",  # Amber / Gold
    "Savings": "\033[38;2;155;89;182m",      # Royal Purple
}

CATEGORY_COLORS = {
    "Mess Food": "2E75B6",
    "Clothes": "548235",
    "Accessories": "BF8F00",
    "Savings": "7030A0",
}


# ==========================================
# LOCALIZATION (i18n) DICTIONARY
# ==========================================
I18N = {
    "en": {
        "greeting": "Hi",
        "title": "REMAINING BUDGET BREAKDOWN",
        "active": "ACTIVE",
        "branch": "Branch",
        "base": "Base",
        "credited": "Credited",
        "total_alloc": "Total Alloc",
        "spent": "Spent",
        "remaining": "Remaining",
        "total": "TOTAL",
        "categories": {
            "Mess Food": "Mess Food",
            "Clothes": "Clothes",
            "Accessories": "Accessories",
            "Savings": "Savings",
        }
    },
    "hi": {
        "greeting": "नमस्ते",
        "title": "बचे हुए बजट का विवरण",
        "active": "सक्रिय",
        "branch": "शाखा",
        "base": "मूल बजट",
        "credited": "जमा राशि",
        "total_alloc": "कुल आवंटित",
        "spent": "खर्च",
        "remaining": "शेष राशि",
        "total": "कुल योग",
        "categories": {
            "Mess Food": "मेस का खाना",
            "Clothes": "कपड़े",
            "Accessories": "सामग्री/सामान",
            "Savings": "बचत",
        }
    },
    "ja": {
        "greeting": "こんにちは",
        "title": "残余予算内訳",
        "active": "有効",
        "branch": "項目",
        "base": "基本予算",
        "credited": "追加入金",
        "total_alloc": "総割当額",
        "spent": "支出済",
        "remaining": "残金",
        "total": "合計",
        "categories": {
            "Mess Food": "食費 (寮食)",
            "Clothes": "衣服費",
            "Accessories": "備品・小物",
            "Savings": "貯金",
        }
    }
}


# ==========================================
# USER CONFIGURATION & PREFERENCES
# ==========================================
DEFAULT_CONFIG = {
    "currency_symbol": "₹",
    "currency_code": "INR",
    "language": "en",
    "color_theme": "cyber_cyan",
}


def load_user_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                return {**DEFAULT_CONFIG, **cfg}
        except Exception:
            return DEFAULT_CONFIG
    return DEFAULT_CONFIG


def save_user_config(cfg: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=4)


def get_cur_sym() -> str:
    return load_user_config().get("currency_symbol", "₹")


def get_cur_code() -> str:
    return load_user_config().get("currency_code", "INR")


def get_text():
    lang = load_user_config().get("language", "en")
    return I18N.get(lang, I18N["en"])


def show_settings_menu():
    while True:
        cfg = load_user_config()
        sym = cfg.get("currency_symbol", "₹")
        code = cfg.get("currency_code", "INR")
        lang = cfg.get("language", "en")
        theme = cfg.get("color_theme", "cyber_cyan")

        box_width = 65
        title_text = "YOSAN CLI PREFERENCES & SETTINGS"
        t_spaces = box_width - len(title_text)
        t_l = t_spaces // 2
        t_r = t_spaces - t_l

        print(f"\n{C.CYAN}╔" + ("═" * box_width) + f"╗{C.RESET}")
        print(f"{C.CYAN}║{C.RESET}{' ' * t_l}{C.BOLD}{C.WHITE}{title_text}{C.RESET}{' ' * t_r}{C.CYAN}║{C.RESET}")
        print(f"{C.CYAN}╚" + ("═" * box_width) + f"╝{C.RESET}")
        print(f"  {C.CYAN}[1]{C.RESET} Currency Notation    : {C.GREEN}{sym} ({code}){C.RESET}")
        print(f"  {C.CYAN}[2]{C.RESET} Display Language     : {C.YELLOW}{lang.upper()}{C.RESET}")
        print(f"  {C.CYAN}[3]{C.RESET} Terminal Color Theme : {C.PURPLE}{theme}{C.RESET}")
        print(f"  {C.CYAN}[4]{C.RESET} Reset to Defaults")
        print(f"  {C.CYAN}[5]{C.RESET} Back to Terminal {C.GRAY}(or 'b'){C.RESET}")
        print(f"{C.GRAY}" + "─" * (box_width + 2) + f"{C.RESET}")

        ch = input(f"{C.CYAN}Select an option (1-5) [Default 5]: {C.RESET}").strip()

        if ch == "1":
            print(f"\n{C.CYAN}Select Currency Symbol & Code:{C.RESET}")
            print("  [1] ₹ INR (Indian Rupee)")
            print("  [2] $ USD (US Dollar)")
            print("  [3] € EUR (Euro)")
            print("  [4] £ GBP (British Pound)")
            print("  [5] ¥ JPY (Japanese Yen)")
            print("  [6] Custom Symbol")
            c_ch = input(f"{C.CYAN}Choose currency (1-6): {C.RESET}").strip()

            currencies = {
                "1": ("₹", "INR"),
                "2": ("$", "USD"),
                "3": ("€", "EUR"),
                "4": ("£", "GBP"),
                "5": ("¥", "JPY"),
            }
            if c_ch in currencies:
                cfg["currency_symbol"], cfg["currency_code"] = currencies[c_ch]
                save_user_config(cfg)
                print(f"{C.GREEN}✔ Currency updated to {cfg['currency_symbol']} ({cfg['currency_code']}).{C.RESET}")
            elif c_ch == "6":
                cust_sym = input("Enter custom symbol: ").strip()
                cust_code = input("Enter 3-letter currency code: ").strip().upper()
                if cust_sym:
                    cfg["currency_symbol"] = cust_sym
                    cfg["currency_code"] = cust_code or "CUSTOM"
                    save_user_config(cfg)
                    print(f"{C.GREEN}✔ Custom currency saved.{C.RESET}")

        elif ch == "2":
            print(f"\n{C.CYAN}Select Interface Language:{C.RESET}")
            print("  [1] English (EN)")
            print("  [2] Hindi (HI - हिन्दी)")
            print("  [3] Japanese (JA - 日本語)")
            l_ch = input(f"{C.CYAN}Choose language (1-3): {C.RESET}").strip()
            lang_map = {"1": "en", "2": "hi", "3": "ja"}
            if l_ch in lang_map:
                cfg["language"] = lang_map[l_ch]
                save_user_config(cfg)
                print(f"{C.GREEN}✔ Language set to {cfg['language'].upper()}.{C.RESET}")

        elif ch == "3":
            print(f"\n{C.CYAN}Select Visual Theme:{C.RESET}")
            print(f"  [1] {C.CYAN}Cyber Cyan (Default){C.RESET}")
            print(f"  [2] {C.GREEN}Emerald Green{C.RESET}")
            print(f"  [3] {C.PURPLE}Purple Neon{C.RESET}")
            t_ch = input(f"{C.CYAN}Choose theme (1-3): {C.RESET}").strip()
            theme_map = {"1": "cyber_cyan", "2": "emerald_green", "3": "purple_neon"}
            if t_ch in theme_map:
                cfg["color_theme"] = theme_map[t_ch]
                save_user_config(cfg)
                print(f"{C.GREEN}✔ Theme updated to {cfg['color_theme']}.{C.RESET}")

        elif ch == "4":
            save_user_config(DEFAULT_CONFIG)
            print(f"{C.YELLOW}✔ Reset all preferences to default factory configuration.{C.RESET}")

        elif ch in ["5", "", "b", "back", "q", "exit"]:
            break


# ==========================================
# DATA DOWNLOAD & BACKUP PACKAGER
# ==========================================
def download_user_data(target_path: str = None):
    init_db()
    sync_to_excel()

    username = get_current_username()
    downloads_dir = Path(target_path) if target_path else (Path.home() / "Downloads")
    export_folder = downloads_dir / f"yosan_export_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    export_folder.mkdir(parents=True, exist_ok=True)

    files_exported = []

    excel_src = Path(get_user_excel_path())
    if excel_src.exists():
        dest = export_folder / excel_src.name
        shutil.copy2(excel_src, dest)
        files_exported.append(dest)

    db_src = Path(get_user_db_path())
    if db_src.exists():
        dest = export_folder / db_src.name
        shutil.copy2(db_src, dest)
        files_exported.append(dest)

    if REPORTS_DIR.exists():
        reports_dest = export_folder / "reports"
        reports_dest.mkdir(exist_ok=True)
        for report_file in REPORTS_DIR.glob(f"*{username}*"):
            dest = reports_dest / report_file.name
            shutil.copy2(report_file, dest)
            files_exported.append(dest)

    box_width = 85
    header_text = "YOSAN DATA PACKAGE EXPORTED SUCCESSFULLY"
    t_spaces = box_width - len(header_text)
    t_l = t_spaces // 2
    t_r = t_spaces - t_l

    print(f"\n{C.CYAN}╔" + ("═" * box_width) + f"╗{C.RESET}")
    print(f"{C.CYAN}║{C.RESET}{' ' * t_l}{C.BOLD}{C.WHITE}{header_text}{C.RESET}{' ' * t_r}{C.CYAN}║{C.RESET}")
    print(f"{C.CYAN}╚" + ("═" * box_width) + f"╝{C.RESET}")
    print(f"  {C.CYAN}•{C.RESET} Export Location : {C.GREEN}{export_folder}{C.RESET}")
    print(f"  {C.CYAN}•{C.RESET} Files Packaged  : {C.YELLOW}{len(files_exported)} files{C.RESET}")
    print(f"{C.GRAY}" + "─" * (box_width + 2) + f"{C.RESET}")
    for f in files_exported:
        print(f"    ✔ {C.WHITE}{f.name}{C.RESET}")
    print(f"{C.CYAN}" + "═" * (box_width + 2) + f"{C.RESET}\n")

    try:
        os.startfile(export_folder)
    except Exception:
        pass


# ==========================================
# 1. DATABASE & USER PATH ISOLATION
# ==========================================
def get_current_username() -> str:
    session_file = DATA_DIR / ".yosan_session.json"
    if session_file.exists():
        try:
            with open(session_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                uname = data.get("username")
                if uname and uname.strip():
                    return uname.strip()
        except Exception:
            pass
    return "default"


def get_user_db_path() -> str:
    username = get_current_username()
    safe_name = "".join(c for c in username if c.isalnum() or c in ("-", "_")).lower()
    return str(DATA_DIR / f"yosan_{safe_name}.db")


def get_user_excel_path() -> str:
    username = get_current_username()
    safe_name = "".join(c for c in username if c.isalnum() or c in ("-", "_")).lower()
    return str(DATA_DIR / f"budget_book_{safe_name}.xlsx")


def get_db_connection():
    db_file = get_user_db_path()
    conn = sqlite3.connect(db_file)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS months (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_code TEXT UNIQUE NOT NULL,
                month_name TEXT NOT NULL,
                total_budget REAL NOT NULL,
                is_active INTEGER DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS allocations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_id INTEGER NOT NULL,
                branch TEXT NOT NULL,
                allocated_amount REAL NOT NULL,
                base_amount REAL DEFAULT 0.0,
                FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE,
                UNIQUE(month_id, branch)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_id INTEGER NOT NULL,
                branch TEXT NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL,
                timestamp TEXT NOT NULL,
                FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS topups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_id INTEGER NOT NULL,
                branch TEXT NOT NULL,
                amount REAL NOT NULL,
                description TEXT DEFAULT 'Budget Top-up',
                timestamp TEXT NOT NULL,
                FOREIGN KEY (month_id) REFERENCES months(id) ON DELETE CASCADE
            )
        """)

        cursor.execute("PRAGMA table_info(allocations)")
        columns_alloc = [row["name"] for row in cursor.fetchall()]
        if "base_amount" not in columns_alloc:
            cursor.execute("ALTER TABLE allocations ADD COLUMN base_amount REAL DEFAULT 0.0")
            cursor.execute("UPDATE allocations SET base_amount = allocated_amount WHERE base_amount = 0.0 OR base_amount IS NULL")

        cursor.execute("PRAGMA table_info(topups)")
        columns_topup = [row["name"] for row in cursor.fetchall()]
        if "description" not in columns_topup:
            cursor.execute("ALTER TABLE topups ADD COLUMN description TEXT DEFAULT 'Budget Top-up'")

        conn.commit()


def get_active_month():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM months WHERE is_active = 1 LIMIT 1")
        return cursor.fetchone()


# ==========================================
# 2. EXCEL SYNC
# ==========================================
def create_borders(thin_color="D9D9D9"):
    return Border(
        left=Side(style="thin", color=thin_color),
        right=Side(style="thin", color=thin_color),
        top=Side(style="thin", color=thin_color),
        bottom=Side(style="thin", color=thin_color),
    )


def sync_to_excel():
    active_m = get_active_month()
    if not active_m:
        return

    sym = get_cur_sym()
    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Summary"
    ws_dash.append([
        "Category",
        f"Base Budget ({sym})",
        f"Credited ({sym})",
        f"Total Allocated ({sym})",
        f"Spent ({sym})",
        f"Remaining ({sym})",
    ])

    dash_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    dash_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws_dash[1]:
        cell.fill = dash_fill
        cell.font = dash_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = create_borders()

    with get_db_connection() as conn:
        cursor = conn.cursor()
        row_idx = 2
        grand_base = 0.0
        grand_credit = 0.0
        grand_alloc = 0.0
        grand_spent = 0.0

        for cat_name in CATEGORIES.values():
            cursor.execute(
                "SELECT allocated_amount, base_amount FROM allocations WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            alloc_row = cursor.fetchone()
            alloc_amt = alloc_row["allocated_amount"] if alloc_row else 0.0
            base_amt = (
                alloc_row["base_amount"]
                if (alloc_row and alloc_row["base_amount"] is not None and alloc_row["base_amount"] > 0)
                else alloc_amt
            )

            cursor.execute(
                "SELECT SUM(amount) AS total FROM topups WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            topup_row = cursor.fetchone()
            credited_amt = topup_row["total"] if topup_row["total"] else 0.0

            cursor.execute(
                "SELECT SUM(amount) AS total FROM transactions WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            spent_row = cursor.fetchone()
            spent_amt = spent_row["total"] if spent_row["total"] else 0.0

            remaining = alloc_amt - spent_amt
            grand_base += base_amt
            grand_credit += credited_amt
            grand_alloc += alloc_amt
            grand_spent += spent_amt

            ws_dash.append([
                cat_name,
                base_amt,
                credited_amt,
                alloc_amt,
                spent_amt,
                remaining,
            ])
            for c in ws_dash[row_idx]:
                c.font = Font(name="Calibri", size=10)
                c.border = create_borders("E0E0E0")
            row_idx += 1

            ws_cat = wb.create_sheet(title=cat_name)
            ws_cat.append(["Timestamp", "Type", "Description", f"Amount ({sym})"])
            color = CATEGORY_COLORS[cat_name]
            h_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in ws_cat[1]:
                cell.fill = h_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = create_borders()
            ws_cat.row_dimensions[1].height = 24

            cursor.execute(
                """
                SELECT timestamp, 'Expense' AS type, description, amount FROM transactions WHERE month_id = ? AND branch = ?
                UNION ALL
                SELECT timestamp, 'Credit' AS type, description, amount FROM topups WHERE month_id = ? AND branch = ?
                ORDER BY timestamp ASC
                """,
                (active_m["id"], cat_name, active_m["id"], cat_name),
            )
            for tx in cursor.fetchall():
                ws_cat.append([
                    tx["timestamp"],
                    tx["type"],
                    tx["description"],
                    tx["amount"],
                ])
                last_r = ws_cat.max_row
                ws_cat.cell(row=last_r, column=1).alignment = Alignment(horizontal="center")
                ws_cat.cell(row=last_r, column=2).alignment = Alignment(horizontal="center")
                ws_cat.cell(row=last_r, column=4).alignment = Alignment(horizontal="right")
                for cell in ws_cat[last_r]:
                    cell.border = create_borders("E0E0E0")

            for col in ws_cat.columns:
                max_l = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_cat.column_dimensions[col_letter].width = max(max_l + 5, 14)

        ws_dash.append([
            "TOTAL",
            grand_base,
            grand_credit,
            grand_alloc,
            grand_spent,
            grand_alloc - grand_spent,
        ])
        for c in ws_dash[row_idx]:
            c.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
            c.border = create_borders()

        for col in ws_dash.columns:
            max_l = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_dash.column_dimensions[col_letter].width = max(max_l + 5, 14)

    target_excel_path = get_user_excel_path()
    try:
        wb.save(target_excel_path)
    except PermissionError:
        pass


# ==========================================
# 3. REPORT EXPORTERS (TXT & PDF)
# ==========================================
def generate_txt_report(target_m, branch_data, grand_base, grand_credit, grand_alloc, grand_spent):
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    username = get_current_username()
    file_path = REPORTS_DIR / f"budget_report_{username}_{target_m['month_code']}.txt"
    code = get_cur_code()

    header_title = (
        f"MONTH-END STATEMENT: {target_m['month_name'].upper()}"
        if target_m["is_active"] == 0
        else f"INTERIM STATEMENT (IN PROGRESS): {target_m['month_name'].upper()}"
    )

    lines = [
        "=" * 98,
        f"                {header_title} [{target_m['month_code']}]",
        "=" * 98,
        f" User           : {username}",
        f" Status         : {'BURNED IN / ARCHIVED (FINAL)' if target_m['is_active'] == 0 else 'ACTIVE / ONGOING'}",
        f" Generated on   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f" Base Budget    : {code} {grand_base:.2f}",
        f" Credited (+)   : {code} {grand_credit:.2f}",
        f" Total Budget   : {code} {grand_alloc:.2f}",
        f" Total Spent    : {code} {grand_spent:.2f}",
        f" Net Balance    : {code} {(grand_alloc - grand_spent):.2f}",
        "=" * 98 + "\n",
    ]

    for cat_name, b_info in branch_data.items():
        base = b_info["base"]
        credit = b_info["credit"]
        alloc = b_info["alloc"]
        spent = b_info["spent"]
        rem = alloc - spent
        lines.append(f"BRANCH: {cat_name.upper()}")
        lines.append(f"Base: {code} {base:.2f} | Credited: +{code} {credit:.2f} | Total: {code} {alloc:.2f} | Spent: {code} {spent:.2f} | Remaining: {code} {rem:.2f}")
        lines.append("-" * 98)
        lines.append(f"{'#':<4} | {'Timestamp':<20} | {'Description':<38} | {f'Debit ({code})':>13} | {f'Credit ({code})':>13}")
        lines.append("-" * 98)
        if b_info["ledger"]:
            for idx, tx in enumerate(b_info["ledger"], start=1):
                debit_str = f"{tx['amount']:>13.2f}" if tx["type"] == "Expense" else f"{'-':>13}"
                credit_str = f"{tx['amount']:>13.2f}" if tx["type"] == "Credit" else f"{'-':>13}"
                lines.append(f"{idx:<4} | {tx['timestamp']:<20} | {tx['description'][:38]:<38} | {debit_str} | {credit_str}")
        else:
            lines.append("     No transactions or top-ups recorded for this branch.")
        lines.append("-" * 98 + "\n")

    with open(file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return file_path


def generate_pdf_report(target_m, branch_data, grand_base, grand_credit, grand_alloc, grand_spent):
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    username = get_current_username()
    file_path = REPORTS_DIR / f"budget_report_{username}_{target_m['month_code']}.pdf"
    code = get_cur_code()

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        rightMargin=32,
        leftMargin=32,
        topMargin=32,
        bottomMargin=32,
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.black,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.black,
    )
    section_head = ParagraphStyle(
        "SectionHead",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.black,
    )
    cell_bold_white = ParagraphStyle(
        "CellBoldWhite",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=10,
        textColor=colors.white,
    )
    cell_bold = ParagraphStyle(
        "CellBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=10,
    )
    cell_regular = ParagraphStyle(
        "CellRegular",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
    )

    if target_m["is_active"] == 0:
        main_header = f"MONTH-END STATEMENT: {target_m['month_name'].upper()}"
        status_label = "BURNED IN / ARCHIVED (FINAL STATEMENT)"
    else:
        main_header = f"INTERIM STATEMENT: {target_m['month_name'].upper()}"
        status_label = "ACTIVE BUDGET CYCLE (IN PROGRESS)"

    story.append(Paragraph(main_header, title_style))
    story.append(Spacer(1, 4))
    story.append(
        Paragraph(
            f"User: <b>{username}</b> &nbsp;|&nbsp; Code: <b>{target_m['month_code']}</b> &nbsp;|&nbsp; Status: <b>{status_label}</b> &nbsp;|&nbsp; Generated: <b>{datetime.now().strftime('%d %b %Y, %H:%M')}</b>",
            meta_style,
        )
    )
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.black, spaceAfter=14))

    summary_data = [
        [
            Paragraph("Category", cell_bold_white),
            Paragraph("Base Alloc", cell_bold_white),
            Paragraph("Credited (+)", cell_bold_white),
            Paragraph("Spent Amount", cell_bold_white),
            Paragraph("Remaining Balance", cell_bold_white),
        ]
    ]

    for cat_name, b_info in branch_data.items():
        rem = b_info["alloc"] - b_info["spent"]
        summary_data.append([
            Paragraph(cat_name, cell_regular),
            f"{code} {b_info['base']:.2f}",
            f"+{code} {b_info['credit']:.2f}" if b_info["credit"] > 0 else f"{code} 0.00",
            f"{code} {b_info['spent']:.2f}",
            f"{code} {rem:.2f}",
        ])

    summary_data.append([
        Paragraph("TOTAL", cell_bold),
        f"{code} {grand_base:.2f}",
        f"+{code} {grand_credit:.2f}",
        f"{code} {grand_spent:.2f}",
        f"{code} {(grand_alloc - grand_spent):.2f}",
    ])

    t_summary = Table(summary_data, colWidths=[135, 95, 95, 100, 105])
    t_summary.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("LINEBELOW", (0, -1), (-1, -1), 1.5, colors.black),
        ])
    )

    story.append(Paragraph("EXECUTIVE SUMMARY", section_head))
    story.append(Spacer(1, 6))
    story.append(t_summary)
    story.append(Spacer(1, 18))

    story.append(Paragraph("DETAILED TRANSACTION LEDGER", section_head))
    story.append(Spacer(1, 8))

    for cat_name, b_info in branch_data.items():
        alloc = b_info["alloc"]
        spent = b_info["spent"]
        credit = b_info["credit"]
        rem = alloc - spent

        credit_str = f" | Credited: +{code} {credit:.2f}" if credit > 0 else ""
        story.append(
            Paragraph(
                f"<b>{cat_name.upper()}</b> &nbsp;—&nbsp; Effective Budget: {code} {alloc:.2f}{credit_str} | Spent: {code} {spent:.2f} | Balance: {code} {rem:.2f}",
                meta_style,
            )
        )
        story.append(Spacer(1, 4))

        branch_table_data = [
            [
                Paragraph("#", cell_bold),
                Paragraph("Timestamp", cell_bold),
                Paragraph("Description", cell_bold),
                Paragraph(f"Debit ({code})", cell_bold),
                Paragraph(f"Credit ({code})", cell_bold),
            ]
        ]

        if b_info["ledger"]:
            for idx, tx in enumerate(b_info["ledger"], start=1):
                debit_val = f"{code} {tx['amount']:.2f}" if tx["type"] == "Expense" else "-"
                credit_val = f"+{code} {tx['amount']:.2f}" if tx["type"] == "Credit" else "-"
                branch_table_data.append([
                    Paragraph(str(idx), cell_regular),
                    Paragraph(tx["timestamp"], cell_regular),
                    Paragraph(tx["description"], cell_regular),
                    debit_val,
                    credit_val,
                ])
        else:
            branch_table_data.append([
                Paragraph("-", cell_regular),
                Paragraph("No records", cell_regular),
                Paragraph("-", cell_regular),
                "-",
                "-",
            ])

        t_branch = Table(branch_table_data, colWidths=[25, 125, 190, 95, 95])
        t_branch.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEAEA")),
                ("ALIGN", (3, 0), (4, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ])
        )
        story.append(t_branch)
        story.append(Spacer(1, 12))

    doc.build(story)
    return file_path


def export_monthly_report(month_code: str = None):
    init_db()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        if month_code:
            cursor.execute("SELECT * FROM months WHERE month_code = ?", (month_code,))
            target_m = cursor.fetchone()
        else:
            active_m = get_active_month()
            if active_m:
                target_m = active_m
            else:
                cursor.execute("SELECT * FROM months ORDER BY id DESC LIMIT 1")
                target_m = cursor.fetchone()

    if not target_m:
        print(f"\n{C.RED}❌ No budget data found to generate reports.{C.RESET}")
        return

    month_id = target_m["id"]
    branch_data = {}
    grand_base = 0.0
    grand_credit = 0.0
    grand_alloc = 0.0
    grand_spent = 0.0

    with get_db_connection() as conn:
        cursor = conn.cursor()
        for cat_name in CATEGORIES.values():
            cursor.execute(
                "SELECT allocated_amount, base_amount FROM allocations WHERE month_id = ? AND branch = ?",
                (month_id, cat_name),
            )
            alloc_row = cursor.fetchone()
            alloc_amt = alloc_row["allocated_amount"] if alloc_row else 0.0
            base_amt = (
                alloc_row["base_amount"]
                if (alloc_row and alloc_row["base_amount"] is not None and alloc_row["base_amount"] > 0)
                else alloc_amt
            )

            cursor.execute(
                "SELECT SUM(amount) AS total FROM topups WHERE month_id = ? AND branch = ?",
                (month_id, cat_name),
            )
            topup_row = cursor.fetchone()
            credit_amt = topup_row["total"] if topup_row["total"] else 0.0

            cursor.execute(
                "SELECT SUM(amount) AS total FROM transactions WHERE month_id = ? AND branch = ?",
                (month_id, cat_name),
            )
            spent_row = cursor.fetchone()
            cat_spent = spent_row["total"] if spent_row["total"] else 0.0

            cursor.execute(
                """
                SELECT timestamp, 'Expense' AS type, description, amount FROM transactions WHERE month_id = ? AND branch = ?
                UNION ALL
                SELECT timestamp, 'Credit' AS type, description, amount FROM topups WHERE month_id = ? AND branch = ?
                ORDER BY timestamp ASC
                """,
                (month_id, cat_name, month_id, cat_name),
            )
            combined_ledger = cursor.fetchall()

            branch_data[cat_name] = {
                "base": base_amt,
                "credit": credit_amt,
                "alloc": alloc_amt,
                "spent": cat_spent,
                "ledger": combined_ledger,
            }
            grand_base += base_amt
            grand_credit += credit_amt
            grand_alloc += alloc_amt
            grand_spent += cat_spent

    txt_file = generate_txt_report(target_m, branch_data, grand_base, grand_credit, grand_alloc, grand_spent)
    pdf_file = generate_pdf_report(target_m, branch_data, grand_base, grand_credit, grand_alloc, grand_spent)

    print(f"\n{C.CYAN}═════════════════════════════════════════════════════════════════{C.RESET}")
    print(f" {C.BOLD}{C.GREEN}REPORT EXPORTED FOR: {target_m['month_name']} [{target_m['month_code']}]{C.RESET}")
    print(f"{C.CYAN}═════════════════════════════════════════════════════════════════{C.RESET}")
    print(f" {C.CYAN}•{C.RESET} Text Report : {C.YELLOW}{txt_file}{C.RESET}")
    print(f" {C.CYAN}•{C.RESET} PDF Report  : {C.YELLOW}{pdf_file}{C.RESET}")
    print(f"{C.CYAN}═════════════════════════════════════════════════════════════════{C.RESET}\n")

    try:
        os.startfile(pdf_file)
    except Exception:
        pass


# ==========================================
# 4. JUJUTSU MANUAL GENERATOR
# ==========================================
def generate_jujutsu_manual():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    file_path = REPORTS_DIR / "yosan_command_manual.pdf"

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.black,
    )
    subtitle_style = ParagraphStyle(
        "DocSub",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.black,
    )
    section_title = ParagraphStyle(
        "SecTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=16,
        textColor=colors.black,
    )
    cell_head = ParagraphStyle(
        "CellHead",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.white,
    )
    cmd_style = ParagraphStyle(
        "CmdStyle",
        parent=styles["Normal"],
        fontName="Courier-Bold",
        fontSize=8.5,
        leading=11,
    )
    desc_style = ParagraphStyle(
        "DescStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
    )

    story.append(Paragraph("YOSAN CLI — MASTER COMMAND MANUAL", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph("System Specification, Operational Flags & Continuous Entry Guide (Jujutsu Edition)", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.black, spaceAfter=14))

    commands = [
        ("yosan", "Displays the live Remaining Budget breakdown and status of the current active month."),
        ("yosan -t, --transactions", "Prints itemized transaction ledger for the active cycle or historical month ('-d MMYYYY')."),
        ("yosan switch [flag]", "Enters continuous entry mode for a category branch. Type 'yosan -juubun' to exit."),
        ("yosan switch [flag] -d \"...\" -val X", "One-line quick entry to log an expense directly to cloud/local ledger."),
        ("yosan -u [flag] [amt] -d \"...\"", "Credits / Adds money to a branch budget and increases overall monthly allowance."),
        ("yosan -new", "Launches Multi-Step Wizard with live auto-fill and boundary caps ('p'=prev, 'b'=abort)."),
        ("yosan -p, yosan profile", "Opens profile dashboard (view account info, change password, delete account)."),
        ("yosan -set, yosan settings", "Opens Preferences Menu to configure Currency Symbol (₹, $, €, £, ¥), Language, and Theme."),
        ("yosan -dl, yosan download", "Packages and downloads database, Excel spreadsheets, and reports to Downloads folder."),
        ("yosan -burn", "Permanently finalizes and locks the active budget cycle into Read-Only mode."),
        ("yosan -report [code]", "Exports formatted Text and styled A4 PDF financial statements for any month."),
        ("yosan -jujutsu", "Generates and opens this Command Manual PDF documentation (Accessible logged out)."),
        ("yosan peek", "Lists all historical budget cycles stored in the active user database."),
        ("yosan peek -d [MMYYYY]", "Inspects read-only ledger and status breakdown of a specific historical month."),
        ("yosan -s", "Displays total expenditure summary across all branch categories."),
        ("yosan -o", "Syncs SQLite records and launches the user-specific budget book in Excel."),
        ("yosan -logout", "Clears the active cloud token session and requires re-authentication."),
        ("yosan -delete-account", "Permanently deletes account credentials, cloud records, and local ledger files."),
    ]

    manual_data = [[Paragraph("Command / Syntax", cell_head), Paragraph("Description & Execution Behavior", cell_head)]]
    for cmd, desc in commands:
        manual_data.append([Paragraph(cmd, cmd_style), Paragraph(desc, desc_style)])

    t_manual = Table(manual_data, colWidths=[185, 340])
    t_manual.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("TOPPADDING", (0, 0), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )

    story.append(Paragraph("CLI REFERENCE MATRIX", section_title))
    story.append(Spacer(1, 6))
    story.append(t_manual)
    story.append(Spacer(1, 14))

    branch_data = [
        [Paragraph("Flag", cell_head), Paragraph("Category Name", cell_head), Paragraph("Default Scope", cell_head)],
        [Paragraph("-m, --mess", cmd_style), Paragraph("Mess Food", desc_style), Paragraph("Daily dining, breakfast, canteen snacks, meals", desc_style)],
        [Paragraph("-c, --clothes", cmd_style), Paragraph("Clothes", desc_style), Paragraph("Apparel, footwear, tailoring, laundry", desc_style)],
        [Paragraph("-a, --accessories", cmd_style), Paragraph("Accessories", desc_style), Paragraph("Electronics, grooming, hardware components, stationery", desc_style)],
        [Paragraph("-v, --savings", cmd_style), Paragraph("Savings", desc_style), Paragraph("Emergency reserve, long-term deposits, investment pool", desc_style)],
    ]

    t_branch = Table(branch_data, colWidths=[110, 130, 285])
    t_branch.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    story.append(Paragraph("BRANCH CATEGORY FLAGS", section_title))
    story.append(Spacer(1, 6))
    story.append(t_branch)
    story.append(Spacer(1, 14))

    story.append(Paragraph("OPERATIONAL FEATURES & SECURITY ARCHITECTURE", section_title))
    story.append(Spacer(1, 4))
    story.append(
        Paragraph(
            "• <b>Smart Wizard ('yosan -new')</b>: Features live boundary protection preventing entry over the maximum remaining limit, auto-fills the remaining balance on the final branch, and renders the live dashboard upon creation.<br/>"
            "• <b>Transaction Inspection ('yosan -t [-d MMYYYY]')</b>: Isolates the itemized debit/credit logs so the primary dashboard remains clean and focused.<br/>"
            "• <b>Continuous Logging ('yosan switch -m')</b>: Enter sequential items without repeating command names. Exit with <b>'yosan -juubun'</b> or <b>'exit'</b>.<br/>"
            "• <b>Custom Preferences ('yosan -set')</b>: Personalize currency signs across your dashboards, terminal tables, and exported files.<br/>"
            "• <b>Cloud Security & Authentication</b>: 1.0s debounced username/email validation, real-time password strength checking (8+ chars, uppercase, digit, symbol), real-time confirmation match badges, and PBKDF2 600k hashing.",
            subtitle_style,
        )
    )

    doc.build(story)
    print(f"\n{C.PURPLE}═════════════════════════════════════════════════════════════════{C.RESET}")
    print(f" {C.BOLD}{C.WHITE}📖 YOSAN JUJUTSU MANUAL GENERATED SUCCESSFULLY{C.RESET}")
    print(f"{C.PURPLE}═════════════════════════════════════════════════════════════════{C.RESET}")
    print(f" {C.CYAN}•{C.RESET} PDF Manual : {C.YELLOW}{file_path}{C.RESET}")
    print(f"{C.PURPLE}═════════════════════════════════════════════════════════════════{C.RESET}\n")

    try:
        os.startfile(file_path)
    except Exception:
        pass


# ==========================================
# 5. CLI CORE FUNCTIONS
# ==========================================
def validate_month_code(code: str):
    if len(code) != 6 or not code.isdigit():
        return False, "Input must be exactly 6 digits in MMYYYY format (e.g., 082026)."
    month_num, year_num = int(code[:2]), int(code[2:])
    if month_num < 1 or month_num > 12:
        return False, f"Invalid month '{code[:2]}'. Must be between 01 and 12."
    if year_num < 2000 or year_num > 2100:
        return False, f"Invalid year '{year_num}'. Must be between 2000 and 2100."
    return True, f"{calendar.month_name[month_num]} {year_num}"


def create_new_budget():
    init_db()
    sym = get_cur_sym()
    print(f"\n{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
    print(f"        {C.BOLD}{C.WHITE}INITIALIZE NEW MONTHLY BUDGET WIZARD{C.RESET}")
    print(f"  {C.GRAY}• Type 'p' / 'prev' at any step to return to previous step{C.RESET}")
    print(f"  {C.GRAY}• Type 'b' / 'cancel' at any step to abort & exit{C.RESET}")
    print(f"{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")

    active_m = get_active_month()
    if active_m:
        print(f"{C.YELLOW}⚠️  An active budget is currently running for: {active_m['month_name']} [{active_m['month_code']}]{C.RESET}")
        burn_confirm = input(f"Are you sure you want to finalize '{active_m['month_name']}' table? (Y/N/B): ").strip().upper()

        if burn_confirm in ["B", "BACK", "CANCEL"]:
            print(f"{C.GRAY}↩ Creation canceled. Returning to terminal.{C.RESET}\n")
            return
        if burn_confirm != "Y":
            print(f"{C.RED}❌ Creation aborted. Current active table remains untouched.{C.RESET}\n")
            return

        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE months SET is_active = 0 WHERE id = ?", (active_m["id"],))
            conn.commit()

        print(f"{C.YELLOW}🔒 '{active_m['month_name']}' has been permanently archived (Read-Only).{C.RESET}")

    step = 1
    raw_code = ""
    month_display = ""
    total_income = 0.0
    branch_allocations = {}

    while True:
        # STEP 1: Month Code
        if step == 1:
            raw_input = input(f"\n{C.CYAN}[Step 1/4]{C.RESET} Enter NEW Month Code (MMYYYY, e.g. 102026) [b=cancel]: ").strip()
            if raw_input.lower() in ["b", "back", "cancel"]:
                print(f"{C.GRAY}↩ Budget creation canceled.{C.RESET}\n")
                return

            is_valid, result = validate_month_code(raw_input)
            if not is_valid:
                print(f"{C.RED}❌ {result}{C.RESET}")
                continue

            month_display = result
            raw_code = raw_input

            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, is_active FROM months WHERE month_code = ?", (raw_code,))
                existing_month = cursor.fetchone()

            if existing_month:
                print(f"\n{C.YELLOW}⚠️  The budget for {month_display} ({raw_code}) already exists in your database!{C.RESET}")
                if existing_month["is_active"] == 1:
                    print(f"👉 {C.GREEN}{month_display} is already active.{C.RESET}\n")
                else:
                    switch_choice = input(f"Reactivate and switch working budget to {month_display}? (Y/N): ").strip().upper()
                    if switch_choice == "Y":
                        with get_db_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute("UPDATE months SET is_active = 0")
                            cursor.execute("UPDATE months SET is_active = 1 WHERE month_code = ?", (raw_code,))
                            conn.commit()
                        sync_to_excel()
                        print(f"{C.GREEN}✅ Reactivated budget {month_display} successfully!{C.RESET}\n")
                return

            print(f" {C.GREEN}Month verified: {month_display}{C.RESET}")
            step = 2

        # STEP 2: Total Budget Amount
        elif step == 2:
            raw_input = input(f"\n{C.CYAN}[Step 2/4]{C.RESET} Enter Total Budget for {C.BOLD}{month_display}{C.RESET} ({sym}) [p=prev step, b=cancel]: ").strip()
            if raw_input.lower() in ["p", "prev", "previous"]:
                print(f"{C.GRAY}↩ Going back to Step 1 (Month Code)...{C.RESET}")
                step = 1
                continue
            if raw_input.lower() in ["b", "back", "cancel"]:
                print(f"{C.GRAY}↩ Budget creation canceled.{C.RESET}\n")
                return

            try:
                income_val = float(raw_input.replace(",", "").replace(sym, "").replace("₹", ""))
                if income_val <= 0:
                    print(f"{C.RED}❌ Total budget must be greater than 0.{C.RESET}")
                    continue
                total_income = income_val
                step = 3
            except ValueError:
                print(f"{C.RED}❌ Invalid input. Please enter a valid numeric amount.{C.RESET}")

        # STEP 3: Assign Branch Budgets
        elif step == 3:
            print(f"\n{C.CYAN}[Step 3/4]{C.RESET} Assign Branch Budgets for {C.BOLD}{month_display}{C.RESET} (Total: {C.GREEN}{sym}{total_income:,.2f}{C.RESET})")
            print(f"  {C.CYAN}[1]{C.RESET} Auto-divide by Percentage Ratio (e.g., 40%, 20%, 15%, 25%)")
            print(f"  {C.CYAN}[2]{C.RESET} Manual Entry (Enter absolute {sym} amounts per branch)")
            print(f"  {C.CYAN}[p]{C.RESET} Previous Step (Go back to change Total Budget)")
            print(f"  {C.CYAN}[b]{C.RESET} Cancel & Exit")

            choice = input(f"{C.CYAN}Select option (1, 2, p, b): {C.RESET}").strip().lower()
            if choice in ["p", "prev", "previous"]:
                print(f"{C.GRAY}↩ Going back to Step 2 (Total Budget Amount)...{C.RESET}")
                step = 2
                continue
            if choice in ["b", "back", "cancel"]:
                print(f"{C.GRAY}↩ Budget creation canceled.{C.RESET}\n")
                return

            # --- OPTION 1: PERCENTAGE ALLOCATION (With Immediate Cap Verification) ---
            if choice == "1":
                print(f"\n{C.GRAY}Enter percentage for each branch (Total must = 100%, 'p'=back):{C.RESET}")
                ratios = {}
                tot_p = 0.0
                step_canceled = False
                category_items = list(CATEGORIES.values())

                for idx, cat_name in enumerate(category_items):
                    cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
                    is_last = (idx == len(category_items) - 1)
                    remaining_p = max(0.0, round(100.0 - tot_p, 2))

                    while True:
                        if is_last:
                            prompt_str = f"  • Percentage for [{cat_color}{cat_name:<11}{C.RESET}] (%): {remaining_p:g} (Auto-filled, Enter to accept): "
                        else:
                            prompt_str = f"  • Percentage for [{cat_color}{cat_name:<11}{C.RESET}] (%) [Max {remaining_p:g}%]: "

                        p_in = input(prompt_str).strip()

                        if p_in.lower() in ["p", "prev", "previous", "b", "back"]:
                            step_canceled = True
                            break

                        if is_last and p_in == "":
                            p = remaining_p
                        else:
                            try:
                                p = float(p_in)
                            except ValueError:
                                print(f"    {C.RED}❌ Numeric percentage required.{C.RESET}")
                                continue

                        if p < 0:
                            print(f"    {C.RED}❌ Value cannot be negative.{C.RESET}")
                            continue

                        if p > remaining_p:
                            print(f"    {C.RED}❌ Value exceeds remaining limit! Maximum allowed is {remaining_p:g}%.{C.RESET}")
                            continue

                        ratios[cat_name] = p
                        tot_p += p
                        break

                    if step_canceled:
                        break

                if step_canceled:
                    continue

                if round(tot_p, 2) == 100.0:
                    branch_allocations = {
                        cat_name: round((pct / 100.0) * total_income, 2)
                        for cat_name, pct in ratios.items()
                    }
                    step = 4
                else:
                    print(f"{C.RED}❌ Percentages sum to {tot_p:.2f}%, must equal exactly 100%.{C.RESET}")

            # --- OPTION 2: ABSOLUTE AMOUNT ALLOCATION (With Immediate Cap Verification) ---
            elif choice == "2":
                print(f"\n{C.GRAY}Enter allocated {sym} amount for each branch ('p'=back):{C.RESET}")
                tot_m = 0.0
                alloc_map = {}
                step_canceled = False
                category_items = list(CATEGORIES.values())

                for idx, cat_name in enumerate(category_items):
                    cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
                    is_last = (idx == len(category_items) - 1)
                    remaining_m = max(0.0, round(total_income - tot_m, 2))

                    while True:
                        if is_last:
                            prompt_str = f"  • Allocated for [{cat_color}{cat_name:<11}{C.RESET}] ({sym}): {sym}{remaining_m:,.2f} (Auto-filled, Enter to accept): "
                        else:
                            prompt_str = f"  • Allocated for [{cat_color}{cat_name:<11}{C.RESET}] ({sym}) [Max {sym}{remaining_m:,.2f}]: "

                        val_in = input(prompt_str).strip()

                        if val_in.lower() in ["p", "prev", "previous", "b", "back"]:
                            step_canceled = True
                            break

                        if is_last and val_in == "":
                            val = remaining_m
                        else:
                            try:
                                val = float(val_in.replace(",", "").replace(sym, "").replace("₹", ""))
                            except ValueError:
                                print(f"    {C.RED}❌ Numeric amount required.{C.RESET}")
                                continue

                        if val < 0:
                            print(f"    {C.RED}❌ Value cannot be negative.{C.RESET}")
                            continue

                        if val > remaining_m:
                            print(f"    {C.RED}❌ Value exceeds remaining budget! Maximum allowed is {sym}{remaining_m:,.2f}.{C.RESET}")
                            continue

                        alloc_map[cat_name] = val
                        tot_m += val
                        break

                    if step_canceled:
                        break

                if step_canceled:
                    continue

                if round(tot_m, 2) == round(total_income, 2):
                    branch_allocations = alloc_map
                    step = 4
                else:
                    print(f"{C.RED}❌ Sum {sym}{tot_m:,.2f} does not match total {sym}{total_income:,.2f}.{C.RESET}")
            else:
                print(f"{C.RED}❌ Invalid option. Enter 1, 2, p, or b.{C.RESET}")

        # STEP 4: Review Summary & Confirmation
        elif step == 4:
            print(f"\n{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
            print(f"  {C.BOLD}[Step 4/4] Summary to Lock In for {month_display}:{C.RESET}")
            for name, amt in branch_allocations.items():
                pct = (amt / total_income) * 100.0
                cat_color = CATEGORY_ANSI.get(name, C.WHITE)
                amt_str = f"{sym}{amt:,.2f}"
                print(f"    - {cat_color}{name:<12}{C.RESET}: {C.YELLOW}{amt_str:>16}{C.RESET} {C.GRAY}({pct:>5.1f}%){C.RESET}")
            total_str = f"{sym}{total_income:,.2f}"
            print(f"  {C.BOLD}Total Budget : {C.GREEN}{total_str:>16}{C.RESET}")
            print(f"{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")

            confirm = input(f"Confirm creation of {month_display}? (Y / p=prev / b=cancel): ").strip().upper()
            if confirm in ["P", "PREV", "PREVIOUS"]:
                print(f"{C.GRAY}↩ Going back to Step 3 (Branch Allocations)...{C.RESET}")
                step = 3
                continue
            if confirm in ["B", "BACK", "CANCEL", "N"]:
                print(f"{C.RED}❌ Budget creation canceled.{C.RESET}\n")
                return

            if confirm == "Y":
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("UPDATE months SET is_active = 0")
                    cursor.execute(
                        """
                        INSERT INTO months (month_code, month_name, total_budget, is_active, created_at)
                        VALUES (?, ?, ?, 1, ?)
                        """,
                        (raw_code, month_display, total_income, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    )
                    month_id = cursor.lastrowid
                    for branch_name, amt in branch_allocations.items():
                        cursor.execute(
                            """
                            INSERT INTO allocations (month_id, branch, allocated_amount, base_amount)
                            VALUES (?, ?, ?, ?)
                            """,
                            (month_id, branch_name, amt, amt),
                        )
                    conn.commit()

                sync_to_excel()
                print(f"\n{C.GREEN}{C.BOLD}🎉 Budget for {month_display} is locked in and active!{C.RESET}\n")
                print_remaining_balance()
                return


def update_branch_budget(cat_name: str, add_amount: float = None, desc: str = None):
    init_db()
    active_m = get_active_month()
    if not active_m:
        print(f"\n{C.RED}🔒 ACCESS DENIED: The previous budget has been burned in and archived (Read-Only).{C.RESET}")
        print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to create an active budget first.\n")
        return

    sym = get_cur_sym()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT allocated_amount, base_amount FROM allocations WHERE month_id = ? AND branch = ?",
            (active_m["id"], cat_name),
        )
        row = cursor.fetchone()
        if not row:
            print(f"{C.RED}❌ Branch '{cat_name}' not found.{C.RESET}")
            return

        current_alloc = row["allocated_amount"]

        if add_amount is None:
            cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
            print(f"\n{C.CYAN}--- Add Money / Increase Budget for [{cat_color}{cat_name}{C.RESET}{C.CYAN}] ---{C.RESET}")
            print(f"Current Allocation: {C.YELLOW}{sym}{current_alloc:.2f}{C.RESET}")
            while True:
                try:
                    add_amount = float(input(f"Enter amount to add to [{cat_name}] ({sym}): "))
                    if add_amount <= 0:
                        continue
                    break
                except ValueError:
                    print(f"{C.RED}❌ Numeric value required.{C.RESET}")

        if not desc:
            user_desc = input("Description / Reason (optional - press Enter for default): ").strip()
            desc = user_desc if user_desc else "Budget Top-up"

        new_alloc = current_alloc + add_amount
        new_total = active_m["total_budget"] + add_amount
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute(
            "UPDATE allocations SET allocated_amount = ? WHERE month_id = ? AND branch = ?",
            (new_alloc, active_m["id"], cat_name),
        )
        cursor.execute(
            "UPDATE months SET total_budget = ? WHERE id = ?",
            (new_total, active_m["id"]),
        )
        cursor.execute(
            "INSERT INTO topups (month_id, branch, amount, description, timestamp) VALUES (?, ?, ?, ?, ?)",
            (active_m["id"], cat_name, add_amount, desc, ts),
        )
        conn.commit()

    sync_to_excel()
    cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
    print(f"\n{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
    print(f" 💰 {C.BOLD}Credited Funds to [{cat_color}{cat_name}{C.RESET}{C.BOLD}] ({active_m['month_name']}){C.RESET}")
    print(f"  • Description         : {C.WHITE}{desc}{C.RESET}")
    print(f"  • Previous Allocation : {C.YELLOW}{sym}{current_alloc:>10.2f}{C.RESET}")
    print(f"  • Credited (+)        : {C.GREEN}+{sym}{add_amount:>9.2f}{C.RESET}")
    print(f"  • New Total Allocation: {C.CYAN}{sym}{new_alloc:>10.2f}{C.RESET}")
    print(f"  • New Total Budget    : {C.GREEN}{sym}{new_total:>10.2f}{C.RESET}")
    print(f"{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}\n")


def add_entry(cat_name: str, desc: str, amount: float):
    init_db()
    active_m = get_active_month()
    if not active_m:
        print(f"\n{C.RED}🔒 ACCESS DENIED: Active budget has been burned in (Read-Only).{C.RESET}")
        print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to create an active budget first.\n")
        return

    sym = get_cur_sym()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO transactions (month_id, branch, description, amount, timestamp)
            VALUES (?, ?, ?, ?, ?)
        """,
            (active_m["id"], cat_name, desc, amount, ts),
        )
        conn.commit()

    sync_to_excel()
    cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
    print(f" {C.GREEN}✔{C.RESET} Saved under [{cat_color}{cat_name}{C.RESET}]: {C.WHITE}{desc}{C.RESET} -> {C.RED}{sym}{amount:.2f}{C.RESET}")


def continuous_interactive_entry(cat_name: str):
    active_m = get_active_month()
    if not active_m:
        print(f"\n{C.RED}🔒 ACCESS DENIED: Active budget has been burned in (Read-Only).{C.RESET}")
        print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to create an active budget first.\n")
        return

    sym = get_cur_sym()
    cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
    print(f"\n{C.CYAN}═════════════════════════════════════════════════════════════════{C.RESET}")
    print(f" 🔄 {C.BOLD}CONTINUOUS LOGGING: [{cat_color}{cat_name.upper()}{C.RESET}{C.BOLD}] ({active_m['month_name']}){C.RESET}")
    print(f"    {C.GRAY}Tip: Type 'yosan -juubun' or 'exit' to finish session.{C.RESET}")
    print(f"{C.CYAN}═════════════════════════════════════════════════════════════════{C.RESET}")

    while True:
        desc = input(f"\n[{cat_color}{cat_name}{C.RESET}] Description: ").strip()

        if desc.lower() in ["yosan -juubun", "juubun", "-juubun", "exit", "quit", "q"]:
            print(f"\n{C.GREEN}✅ Finished session for [{cat_name}]. (Juubun / 充分){C.RESET}")
            break

        if not desc:
            print(f"{C.RED}❌ Description cannot be empty.{C.RESET}")
            continue

        while True:
            amount_str = input(f"[{cat_color}{cat_name}{C.RESET}] Amount for '{desc}' ({sym}): ").strip()
            if amount_str.lower() in ["yosan -juubun", "juubun", "-juubun", "exit", "quit", "q"]:
                print(f"\n{C.GREEN}✅ Finished session for [{cat_name}]. (Juubun / 充分){C.RESET}")
                return
            try:
                amount = float(amount_str.replace(",", "").replace(sym, "").replace("₹", ""))
                if amount <= 0:
                    print(f"{C.RED}Amount must be greater than 0.{C.RESET}")
                    continue
                add_entry(cat_name, desc, amount)
                break
            except ValueError:
                print(f"{C.RED}❌ Invalid amount. Enter numeric {sym} value.{C.RESET}")

    print_remaining_balance()


def print_remaining_balance():
    init_db()
    active_m = get_active_month()
    username = get_current_username()
    sym = get_cur_sym()
    t = get_text()

    if not active_m:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM months ORDER BY id DESC LIMIT 1")
            latest_m = cursor.fetchone()

        if not latest_m:
            print(f"\n{t['greeting']} {C.CYAN}{C.BOLD}{username}{C.RESET}!")
            print(f"\n{C.RED}❌ No budget records found. Run '{C.CYAN}yosan -new{C.RESET}{C.RED}' to create one.{C.RESET}")
            return

        print(f"\n{t['greeting']} {C.CYAN}{C.BOLD}{username}{C.RESET}!")
        print(f"\n{C.YELLOW}🔒 ALL BUDGETS ARE BURNED IN (READ-ONLY MODE){C.RESET}")
        print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to initialize a new month's active budget.")
        peek_month_budget(latest_m["month_code"])
        return

    # Total width of table: 101 characters (box_width = 99)
    title_text = f"{t['title']} ({active_m['month_name'].upper()}) [{t['active']}]"
    box_width = 99
    t_spaces = max(0, box_width - len(title_text))
    t_l = t_spaces // 2
    t_r = t_spaces - t_l

    print(f"\n{t['greeting']} {C.CYAN}{C.BOLD}{username}{C.RESET}!")
    print(f"{C.CYAN}╔" + ("═" * box_width) + f"╗{C.RESET}")
    print(f"{C.CYAN}║{C.RESET}{' ' * t_l}{C.BOLD}{C.WHITE}{title_text}{C.RESET}{' ' * t_r}{C.CYAN}║{C.RESET}")
    print(f"{C.CYAN}╚" + ("═" * box_width) + f"╝{C.RESET}")

    # Layout: Branch(13) | Base(15) | Credited(13) | Total Alloc(17) | Spent(15) | Remaining(16)
    print(f"{C.BOLD}{t['branch']:<13} | {t['base']:>15} | {t['credited']:>13} | {t['total_alloc']:>17} | {t['spent']:>15} | {t['remaining']:>16}{C.RESET}")
    print(f"{C.GRAY}" + "─" * 101 + f"{C.RESET}")

    with get_db_connection() as conn:
        cursor = conn.cursor()
        grand_base = 0.0
        grand_credit = 0.0
        grand_alloc = 0.0
        grand_spent = 0.0

        for cat_name in CATEGORIES.values():
            cursor.execute(
                "SELECT allocated_amount, base_amount FROM allocations WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            alloc_row = cursor.fetchone()
            alloc_amt = alloc_row["allocated_amount"] if alloc_row else 0.0
            base_amt = (
                alloc_row["base_amount"]
                if (alloc_row and alloc_row["base_amount"] is not None and alloc_row["base_amount"] > 0)
                else alloc_amt
            )

            cursor.execute(
                "SELECT SUM(amount) AS total FROM topups WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            topup_row = cursor.fetchone()
            credited_amt = topup_row["total"] if topup_row["total"] else 0.0

            cursor.execute(
                "SELECT SUM(amount) AS total FROM transactions WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            spent_row = cursor.fetchone()
            spent_amt = spent_row["total"] if spent_row["total"] else 0.0

            remaining = alloc_amt - spent_amt
            grand_base += base_amt
            grand_credit += credited_amt
            grand_alloc += alloc_amt
            grand_spent += spent_amt

            cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
            display_cat = t["categories"].get(cat_name, cat_name)

            base_str = f"{sym}{base_amt:,.2f}"
            credit_str = f"+{sym}{credited_amt:,.2f}" if credited_amt > 0 else f"{sym}{credited_amt:,.2f}"
            alloc_str = f"{sym}{alloc_amt:,.2f}"
            spent_str = f"{sym}{spent_amt:,.2f}"
            rem_str = f"{sym}{remaining:,.2f}" if remaining >= 0 else f"-{sym}{abs(remaining):,.2f}"

            print(
                f"{cat_color}{display_cat:<13}{C.RESET} | "
                f"{C.WHITE}{base_str:>15}{C.RESET} | "
                f"{C.GREEN if credited_amt > 0 else C.GRAY}{credit_str:>13}{C.RESET} | "
                f"{C.CYAN}{alloc_str:>17}{C.RESET} | "
                f"{C.RED}{spent_str:>15}{C.RESET} | "
                f"{C.GREEN if remaining >= 0 else C.RED}{rem_str:>16}{C.RESET}"
            )

        print(f"{C.GRAY}" + "─" * 101 + f"{C.RESET}")
        total_rem = grand_alloc - grand_spent
        tb_str = f"{sym}{grand_base:,.2f}"
        tc_str = f"+{sym}{grand_credit:,.2f}" if grand_credit > 0 else f"{sym}{grand_credit:,.2f}"
        ta_str = f"{sym}{grand_alloc:,.2f}"
        ts_str = f"{sym}{grand_spent:,.2f}"
        tr_str = f"{sym}{total_rem:,.2f}" if total_rem >= 0 else f"-{sym}{abs(total_rem):,.2f}"

        print(
            f"{C.BOLD}{t['total']:<13}{C.RESET} | "
            f"{C.YELLOW}{tb_str:>15}{C.RESET} | "
            f"{C.GREEN if grand_credit > 0 else C.GRAY}{tc_str:>13}{C.RESET} | "
            f"{C.CYAN}{ta_str:>17}{C.RESET} | "
            f"{C.RED}{ts_str:>15}{C.RESET} | "
            f"{C.GREEN if total_rem >= 0 else C.RED}{C.BOLD}{tr_str:>16}{C.RESET}"
        )
        print(f"{C.CYAN}" + "═" * 101 + f"{C.RESET}\n")


def show_transaction_ledger(month_code: str = None):
    init_db()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        if month_code:
            cursor.execute("SELECT * FROM months WHERE month_code = ?", (month_code,))
            target_m = cursor.fetchone()
        else:
            target_m = get_active_month()

    if not target_m:
        print(f"\n{C.RED}❌ No active budget found. Run 'yosan -new' or provide a month code.{C.RESET}\n")
        return

    sym = get_cur_sym()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT timestamp, 'Expense' AS type, branch, description, amount FROM transactions WHERE month_id = ?
            UNION ALL
            SELECT timestamp, 'Credit' AS type, branch, description, amount FROM topups WHERE month_id = ?
            ORDER BY timestamp ASC
            """,
            (target_m["id"], target_m["id"]),
        )
        tx_list = cursor.fetchall()

    print(f"\n{C.CYAN}═════════════════════════════════════════════════════════════════════════════════════{C.RESET}")
    print(f"  {C.BOLD}TRANSACTION LEDGER: {target_m['month_name'].upper()} [{target_m['month_code']}]{C.RESET}")
    print(f"{C.CYAN}═════════════════════════════════════════════════════════════════════════════════════{C.RESET}")

    if not tx_list:
        print(f"\n{C.GRAY}  ℹ️  No transactions recorded for this month.{C.RESET}\n")
        return

    for idx, tx in enumerate(tx_list, start=1):
        amt_str = f"{C.GREEN}+{sym}{tx['amount']:>9.2f}{C.RESET}" if tx["type"] == "Credit" else f"{C.RED}{sym}{tx['amount']:>9.2f}{C.RESET}"
        cat_color = CATEGORY_ANSI.get(tx["branch"], C.WHITE)
        print(f"  {C.GRAY}{idx:>2}.{C.RESET} [{C.GRAY}{tx['timestamp']}{C.RESET}] [{C.BOLD}{tx['type']:<7}{C.RESET}] [{cat_color}{tx['branch']:<12}{C.RESET}] {tx['description']:<30} -> {amt_str}")

    print(f"{C.CYAN}═════════════════════════════════════════════════════════════════════════════════════{C.RESET}\n")


def print_summary():
    init_db()
    active_m = get_active_month()
    if not active_m:
        print(f"\n{C.RED}❌ No active budget found. Run '{C.CYAN}yosan -new{C.RESET}{C.RED}' first.{C.RESET}")
        return

    sym = get_cur_sym()
    print(f"\n{C.CYAN}═════════════════════════════════════════════{C.RESET}")
    print(f"       {C.BOLD}{C.WHITE}YOSAN SPENDING SUMMARY ({active_m['month_name']}){C.RESET}")
    print(f"{C.CYAN}═════════════════════════════════════════════{C.RESET}")

    with get_db_connection() as conn:
        cursor = conn.cursor()
        grand_total = 0.0
        for cat_name in CATEGORIES.values():
            cursor.execute(
                "SELECT SUM(amount) AS total FROM transactions WHERE month_id = ? AND branch = ?",
                (active_m["id"], cat_name),
            )
            row = cursor.fetchone()
            cat_total = row["total"] if row["total"] else 0.0
            cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
            print(f" * {cat_color}{cat_name:<20}{C.RESET}: {C.RED}{sym}{cat_total:>10.2f}{C.RESET}")
            grand_total += cat_total

        print(f"{C.GRAY}─────────────────────────────────────────────{C.RESET}")
        print(f" {C.BOLD}GRAND TOTAL EXPENDITURE:{C.RESET} {C.RED}{C.BOLD}{sym}{grand_total:>10.2f}{C.RESET}")
        print(f"{C.CYAN}═════════════════════════════════════════════{C.RESET}\n")


def peek_month_budget(month_code: str):
    init_db()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM months WHERE month_code = ?", (month_code,))
        target_m = cursor.fetchone()

    if not target_m:
        print(f"\n{C.RED}❌ No budget record found for Month Code: '{month_code}'.{C.RESET}")
        print(f"{C.GRAY}Tip: Run 'yosan peek' without flags to view all available months.{C.RESET}")
        return

    sym = get_cur_sym()
    active_tag = f" {C.GREEN}(CURRENT ACTIVE){C.RESET}" if target_m["is_active"] == 1 else f" {C.YELLOW}(ARCHIVED - READ ONLY){C.RESET}"
    title_text = f"PEEK HISTORICAL BUDGET: {target_m['month_name'].upper()} [{month_code}]"
    box_width = 99
    t_spaces = max(0, box_width - len(title_text))
    t_l = t_spaces // 2
    t_r = t_spaces - t_l

    print(f"\n{C.CYAN}╔" + ("═" * box_width) + f"╗{C.RESET}")
    print(f"{C.CYAN}║{C.RESET}{' ' * t_l}{C.BOLD}{C.WHITE}{title_text}{C.RESET}{' ' * t_r}{C.CYAN}║{C.RESET}")
    print(f"{C.CYAN}╚" + ("═" * box_width) + f"╝{C.RESET}")
    print(f"Status: {active_tag}")
    print(f"{C.BOLD}{'Branch':<13} | {'Base':>15} | {'Credited':>13} | {'Total Alloc':>17} | {'Spent':>15} | {'Remaining':>16}{C.RESET}")
    print(f"{C.GRAY}" + "─" * 101 + f"{C.RESET}")

    with get_db_connection() as conn:
        cursor = conn.cursor()
        grand_base = 0.0
        grand_credit = 0.0
        grand_alloc = 0.0
        grand_spent = 0.0

        for cat_name in CATEGORIES.values():
            cursor.execute(
                "SELECT allocated_amount, base_amount FROM allocations WHERE month_id = ? AND branch = ?",
                (target_m["id"], cat_name),
            )
            alloc_row = cursor.fetchone()
            alloc_amt = alloc_row["allocated_amount"] if alloc_row else 0.0
            base_amt = (
                alloc_row["base_amount"]
                if (alloc_row and alloc_row["base_amount"] is not None and alloc_row["base_amount"] > 0)
                else alloc_amt
            )

            cursor.execute(
                "SELECT SUM(amount) AS total FROM topups WHERE month_id = ? AND branch = ?",
                (target_m["id"], cat_name),
            )
            topup_row = cursor.fetchone()
            credited_amt = topup_row["total"] if topup_row["total"] else 0.0

            cursor.execute(
                "SELECT SUM(amount) AS total FROM transactions WHERE month_id = ? AND branch = ?",
                (target_m["id"], cat_name),
            )
            stats_row = cursor.fetchone()
            spent_amt = stats_row["total"] if stats_row["total"] else 0.0

            remaining = alloc_amt - spent_amt
            grand_base += base_amt
            grand_credit += credited_amt
            grand_alloc += alloc_amt
            grand_spent += spent_amt

            cat_color = CATEGORY_ANSI.get(cat_name, C.WHITE)
            base_str = f"{sym}{base_amt:,.2f}"
            credit_str = f"+{sym}{credited_amt:,.2f}" if credited_amt > 0 else f"{sym}{credited_amt:,.2f}"
            alloc_str = f"{sym}{alloc_amt:,.2f}"
            spent_str = f"{sym}{spent_amt:,.2f}"
            rem_str = f"{sym}{remaining:,.2f}" if remaining >= 0 else f"-{sym}{abs(remaining):,.2f}"

            print(
                f"{cat_color}{cat_name:<13}{C.RESET} | "
                f"{C.WHITE}{base_str:>15}{C.RESET} | "
                f"{C.GREEN if credited_amt > 0 else C.GRAY}{credit_str:>13}{C.RESET} | "
                f"{C.CYAN}{alloc_str:>17}{C.RESET} | "
                f"{C.RED}{spent_str:>15}{C.RESET} | "
                f"{C.GREEN if remaining >= 0 else C.RED}{rem_str:>16}{C.RESET}"
            )

        print(f"{C.GRAY}" + "─" * 101 + f"{C.RESET}")
        total_rem = grand_alloc - grand_spent
        tb_str = f"{sym}{grand_base:,.2f}"
        tc_str = f"+{sym}{grand_credit:,.2f}" if grand_credit > 0 else f"{sym}{grand_credit:,.2f}"
        ta_str = f"{sym}{grand_alloc:,.2f}"
        ts_str = f"{sym}{grand_spent:,.2f}"
        tr_str = f"{sym}{total_rem:,.2f}" if total_rem >= 0 else f"-{sym}{abs(total_rem):,.2f}"

        print(
            f"{C.BOLD}{'TOTAL':<13}{C.RESET} | "
            f"{C.YELLOW}{tb_str:>15}{C.RESET} | "
            f"{C.GREEN if grand_credit > 0 else C.GRAY}{tc_str:>13}{C.RESET} | "
            f"{C.CYAN}{ta_str:>17}{C.RESET} | "
            f"{C.RED}{ts_str:>15}{C.RESET} | "
            f"{C.GREEN if total_rem >= 0 else C.RED}{C.BOLD}{tr_str:>16}{C.RESET}"
        )
        print(f"{C.CYAN}" + "═" * 101 + f"{C.RESET}\n")


def list_all_available_months():
    init_db()
    sym = get_cur_sym()
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT month_code, month_name, total_budget, is_active FROM months ORDER BY id DESC")
        rows = cursor.fetchall()

    if not rows:
        print(f"\n{C.RED}❌ No months found in database.{C.RESET}")
        return

    print(f"\n{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
    print(f"              {C.BOLD}{C.WHITE}AVAILABLE STORED MONTHS{C.RESET}")
    print(f"{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
    print(f"{C.BOLD}{'Code':<8} | {'Month Name':<18} | {'Budget':>11} | {'Status'}{C.RESET}")
    print(f"{C.GRAY}───────────────────────────────────────────────────────{C.RESET}")
    for r in rows:
        status = f"{C.GREEN}Active{C.RESET}" if r["is_active"] == 1 else f"{C.GRAY}Archived{C.RESET}"
        print(f"{C.YELLOW}{r['month_code']:<8}{C.RESET} | {r['month_name']:<18} | {C.GREEN}{sym}{r['total_budget']:>10.2f}{C.RESET} | {status}")
    print(f"{C.CYAN}═══════════════════════════════════════════════════════{C.RESET}")
    print(f"{C.GRAY}Run: yosan peek -d <code (e.g. 082026)>{C.RESET}\n")


def burn_current_budget():
    init_db()
    active_m = get_active_month()
    if not active_m:
        print(f"\n{C.YELLOW}🔒 There is no active budget running. Everything is already burned in / read-only.{C.RESET}")
        return

    print(f"\n{C.YELLOW}⚠️  Current Active Budget: {active_m['month_name']} [{active_m['month_code']}]{C.RESET}")
    confirm = input(f"Are you sure you want to {C.RED}BURN IN{C.RESET} and lock '{active_m['month_name']}' permanently? (Y/N): ").strip().upper()
    if confirm == "Y":
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("UPDATE months SET is_active = 0 WHERE id = ?", (active_m["id"],))
            conn.commit()
        print(f"{C.GREEN} '{active_m['month_name']}' is now BURNED IN and set to READ-ONLY.{C.RESET}\n")
    else:
        print(f"{C.GRAY}❌ Action canceled. Active budget remains untouched.{C.RESET}\n")


# ==========================================
# 6. MAIN ROUTING
# ==========================================
def main():
    if "-jujutsu" in sys.argv or "--jujutsu" in sys.argv or "jujutsu" in sys.argv:
        generate_jujutsu_manual()
        return

    if "-logout" in sys.argv or "--logout" in sys.argv:
        clear_session()
        return

    if "-delete-account" in sys.argv or "--delete-account" in sys.argv or "-delete" in sys.argv:
        delete_account_flow()
        return

    if not require_login():
        sys.exit(1)

    if "-p" in sys.argv or "--profile" in sys.argv or "profile" in sys.argv:
        show_profile_view()
        return

    if "-set" in sys.argv or "--settings" in sys.argv or "settings" in sys.argv:
        show_settings_menu()
        return

    init_db()

    if "-dl" in sys.argv or "--download" in sys.argv or "download" in sys.argv:
        dest_dir = None
        for arg in sys.argv[1:]:
            if arg not in ["-dl", "--download", "download"] and not arg.startswith("-"):
                dest_dir = arg
                break
        download_user_data(dest_dir)
        return

    # Route: yosan -t [-d MMYYYY / MMYYYY]
    if "-t" in sys.argv or "--transactions" in sys.argv:
        target_code = None
        if "-d" in sys.argv:
            d_idx = sys.argv.index("-d")
            if d_idx + 1 < len(sys.argv):
                target_code = sys.argv[d_idx + 1]
        else:
            for arg in sys.argv[1:]:
                if arg not in ["-t", "--transactions"] and not arg.startswith("-"):
                    target_code = arg
                    break
        show_transaction_ledger(target_code)
        return

    if "-report" in sys.argv or "--report" in sys.argv:
        target_code = None
        if "-d" in sys.argv:
            d_idx = sys.argv.index("-d")
            if d_idx + 1 < len(sys.argv):
                target_code = sys.argv[d_idx + 1]
        else:
            for arg in sys.argv[1:]:
                if arg not in ["-report", "--report"] and not arg.startswith("-"):
                    target_code = arg
                    break
        export_monthly_report(target_code)
        return

    if "-u" in sys.argv or "--update" in sys.argv:
        active_m = get_active_month()
        if not active_m:
            print(f"\n{C.RED}🔒 ACCESS DENIED: Active budget has been burned in (Read-Only).{C.RESET}")
            print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to create an active budget first.\n")
            return

        target_cat = None
        amount_val = None
        desc_val = None
        sym = get_cur_sym()

        if "-d" in sys.argv:
            d_idx = sys.argv.index("-d")
            if d_idx + 1 < len(sys.argv):
                desc_val = sys.argv[d_idx + 1]

        if "-m" in sys.argv:
            target_cat = "Mess Food"
        elif "-c" in sys.argv:
            target_cat = "Clothes"
        elif "-a" in sys.argv:
            target_cat = "Accessories"
        elif "-v" in sys.argv:
            target_cat = "Savings"

        for arg in sys.argv[1:]:
            if arg not in [
                "-u", "--update", "-m", "--mess", "-c", "--clothes",
                "-a", "--accessories", "-v", "--savings", "-d", desc_val,
            ]:
                try:
                    amount_val = float(arg.replace(",", "").replace(sym, "").replace("₹", ""))
                    break
                except ValueError:
                    pass

        if target_cat:
            update_branch_budget(target_cat, amount_val, desc_val)
        else:
            print(f"\n{C.CYAN}Select a branch to add money to:{C.RESET}")
            print(f"  {C.CYAN}[1]{C.RESET} Mess Food (-m)")
            print(f"  {C.CYAN}[2]{C.RESET} Clothes (-c)")
            print(f"  {C.CYAN}[3]{C.RESET} Accessories (-a)")
            print(f"  {C.CYAN}[4]{C.RESET} Savings (-v)")
            ch = input(f"{C.CYAN}Enter choice (1-4): {C.RESET}").strip()
            mapping = {
                "1": "Mess Food",
                "2": "Clothes",
                "3": "Accessories",
                "4": "Savings",
            }
            if ch in mapping:
                update_branch_budget(mapping[ch], amount_val, desc_val)
            else:
                print(f"{C.RED}❌ Invalid selection.{C.RESET}")
        return

    parser = argparse.ArgumentParser(description="Yosan Budget CLI (DB Engine)")
    parser.add_argument("-new", "--new", action="store_true", help="Initialize a new monthly budget")
    parser.add_argument("-burn", "--burn", action="store_true", help="Burn in / finalize the active budget to read-only")
    parser.add_argument("-jujutsu", "--jujutsu", action="store_true", help="Generate Command Reference Manual PDF")
    parser.add_argument("-s", "--summary", action="store_true", help="View sum total balance spent")
    parser.add_argument("-t", "--transactions", nargs="?", const=True, help="View itemized transaction log")
    parser.add_argument("-d", "--date", help="Target Month Code (e.g. 032033)")
    parser.add_argument("-sh", "--show-remaining", action="store_true", help="View remaining branch allowance")
    parser.add_argument("-o", "--open", action="store_true", help="Open budget book in Excel")
    parser.add_argument("-p", "--profile", action="store_true", help="View profile and account settings")
    parser.add_argument("-set", "--settings", action="store_true", help="Open preferences and currency configuration")
    parser.add_argument("-dl", "--download", nargs="?", const=True, help="Download ledgers and reports package")
    parser.add_argument("-logout", "--logout", action="store_true", help="Log out active session")
    parser.add_argument("-delete-account", "--delete-account", action="store_true", help="Permanently delete active account and data")

    subparsers = parser.add_subparsers(dest="subcommand")

    subparsers.add_parser("profile", help="Manage user profile and security settings")
    subparsers.add_parser("settings", help="Configure currency symbol, language, and CLI theme")
    subparsers.add_parser("download", help="Export and download financial ledgers to Downloads folder")
    subparsers.add_parser("jujutsu", help="Open the Command Manual PDF")

    switch_parser = subparsers.add_parser("switch", help="Switch to a specific category")
    switch_parser.add_argument("-m", "--mess", action="store_true", help="Mess Food")
    switch_parser.add_argument("-c", "--clothes", action="store_true", help="Clothes")
    switch_parser.add_argument("-a", "--accessories", action="store_true", help="Accessories")
    switch_parser.add_argument("-v", "--savings", action="store_true", help="Savings")
    switch_parser.add_argument("-d", "--desc", help="Description")
    switch_parser.add_argument("-val", "--amount", type=float, help="Amount")

    peek_parser = subparsers.add_parser("peek", help="Inspect read-only historical budget data")
    peek_parser.add_argument("-d", "--date", help="Month code to peek (e.g., 082026)")

    args = parser.parse_args()

    if args.logout:
        clear_session()
        return

    if args.delete_account:
        delete_account_flow()
        return

    if args.profile or args.subcommand == "profile":
        show_profile_view()
        return

    if args.settings or args.subcommand == "settings":
        show_settings_menu()
        return

    if args.download or args.subcommand == "download":
        dest = args.download if isinstance(args.download, str) else None
        download_user_data(dest)
        return

    if args.jujutsu or args.subcommand == "jujutsu":
        generate_jujutsu_manual()
        return

    if args.burn:
        burn_current_budget()
        return

    if args.new:
        create_new_budget()
        return

    if args.transactions:
        show_transaction_ledger(args.date)
        return

    if args.show_remaining:
        print_remaining_balance()
        return

    if args.summary:
        print_summary()
        return

    if args.open:
        sync_to_excel()
        try:
            os.startfile(get_user_excel_path())
        except Exception:
            pass
        return

    if args.subcommand == "peek":
        if args.date:
            peek_month_budget(args.date)
        else:
            list_all_available_months()
        return

    if args.subcommand == "switch":
        active_m = get_active_month()
        if not active_m:
            print(f"\n{C.RED}🔒 ACCESS DENIED: Active budget has been burned in (Read-Only).{C.RESET}")
            print(f"👉 Run '{C.CYAN}yosan -new{C.RESET}' to create an active budget first.\n")
            return

        target_cat = None
        if args.mess:
            target_cat = "Mess Food"
        elif args.clothes:
            target_cat = "Clothes"
        elif args.accessories:
            target_cat = "Accessories"
        elif args.savings:
            target_cat = "Savings"

        if target_cat:
            if args.desc and args.amount is not None:
                add_entry(target_cat, args.desc, args.amount)
            else:
                continuous_interactive_entry(target_cat)
        else:
            print(f"{C.RED}Please specify a category flag: -m, -c, -a, or -v{C.RESET}")
        return

    print_remaining_balance()


if __name__ == "__main__":
    main()
